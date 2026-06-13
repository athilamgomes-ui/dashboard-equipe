#!/usr/bin/env python3
"""
parse_planilha_pagar.py — lê a planilha "Gestao Compras 2026.xlsx" (baixada do Drive)
e calcula Contas a Pagar por loja, EXCLUINDO pedidos com Status "ATHILA" ou "ENTREGUE"
(pedidos já entregues/pagos não geram contas futuras a pagar).

Determinístico, SEM LLM. Usa openpyxl (data_only=True) → lê os NÚMEROS reais das
células (sem ambiguidade de separador decimal "1,234.56" vs "1.234,56").

Abas: ALTAMIRA, SANTAREM, ITAITUBA (mesmo layout). Colunas:
  A Marcas · B Data Pedido · C Valor Total · D Status · E Forma · F Parcelas
  G Data entrega · H.. = meses (Abril/2026, Maio/2026, ...)

Saída (stdout, JSON), alinhada a uma janela de 5 meses a partir do mês atual:
  {"geradoEm","meses":[5 labels "Mmm/AA"],
   "ALTAMIRA":[5], "SANTAREM":[5], "ITAITUBA":[5]}
ALTAMIRA é o total Altamira (a divisão L1/L4 = /2 é feita no build).

Uso: python3 parse_planilha_pagar.py <arquivo.xlsx> [ANO_INICIAL MES_INICIAL]
  sem args de data → janela = mês corrente .. +4 meses.
"""
import sys, json, datetime
import openpyxl

EXCLUIR = {"ATHILA", "ENTREGUE"}                  # status que NÃO contam como conta a pagar
ABAS = ["ALTAMIRA", "SANTAREM", "ITAITUBA"]
MES_NOME = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
            7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}
MES_ABBR = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",
            7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}

def log(*a): print("[planilha]", *a, file=sys.stderr)

def num(v):
    """Converte célula em float. data_only → já vem número; tolera string formatada."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    s = s.replace("R$", "").replace(" ", "")
    if s in ("-", "—"):
        return 0.0
    # Detecta formato por posição do último separador (decimal)
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):   # BR: 1.234,56 → vírgula é decimal
            s = s.replace(".", "").replace(",", ".")
        else:                              # US: 1,234.56 → ponto é decimal
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0

def janela(ano, mes, n=5):
    """n meses (ano, mes) a partir de (ano, mes)."""
    out = []
    y, m = ano, mes
    for _ in range(n):
        out.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1
    return out

def parse_aba(ws, meses_alvo):
    """Soma colunas mensais das linhas com Status NÃO em EXCLUIR."""
    # Acha a linha de cabeçalho (contém 'Status do Pedido')
    header_row = None
    for r in range(1, min(6, ws.max_row + 1)):
        linha = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any("status do pedido" in x for x in linha):
            header_row = r
            break
    if header_row is None:
        header_row = 1
    # Mapeia "Nome/AAAA" → coluna
    col_de_mes = {}        # (ano,mes) -> col index
    col_status = col_marca = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(header_row, c).value or "").strip()
        hl = h.lower()
        if "status do pedido" in hl:
            col_status = c
        elif hl == "marcas":
            col_marca = c
        else:
            # tenta "Abril/2026"
            for mnum, mnome in MES_NOME.items():
                if hl.startswith(mnome.lower()) and "/" in h:
                    try:
                        ano = int(h.split("/")[1])
                        col_de_mes[(ano, mnum)] = c
                    except (ValueError, IndexError):
                        pass
    if col_status is None:
        col_status = 4
    if col_marca is None:
        col_marca = 1

    somas = []
    incluidas = excluidas = 0
    for (ano, mes) in meses_alvo:
        col = col_de_mes.get((ano, mes))
        if col is None:
            somas.append(0.0)
            continue
        total = 0.0
        for r in range(header_row + 1, ws.max_row + 1):
            marca = str(ws.cell(r, col_marca).value or "").strip()
            if not marca or marca.upper().startswith("TOTAL"):
                continue
            status = str(ws.cell(r, col_status).value or "").strip().upper()
            v = num(ws.cell(r, col).value)
            if v == 0:
                continue
            if status in EXCLUIR:
                if mes == meses_alvo[0][1]:
                    excluidas += 1
                continue
            if mes == meses_alvo[0][1]:
                incluidas += 1
            total += v
        somas.append(round(total))
    return somas, incluidas, excluidas

def main():
    if len(sys.argv) < 2:
        log("uso: parse_planilha_pagar.py <arquivo.xlsx> [ANO MES]")
        sys.exit(2)
    arq = sys.argv[1]
    if len(sys.argv) >= 4:
        ano0, mes0 = int(sys.argv[2]), int(sys.argv[3])
    else:
        hoje = datetime.date.today()
        ano0, mes0 = hoje.year, hoje.month
    meses_alvo = janela(ano0, mes0, 5)
    labels = [f"{MES_ABBR[m]}/{str(y)[2:]}" for (y, m) in meses_alvo]

    wb = openpyxl.load_workbook(arq, data_only=True)
    out = {"geradoEm": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
           "meses": labels}
    for aba in ABAS:
        if aba not in wb.sheetnames:
            log(f"AVISO: aba '{aba}' ausente")
            out[aba] = [0, 0, 0, 0, 0]
            continue
        somas, inc, exc = parse_aba(wb[aba], meses_alvo)
        out[aba] = somas
        log(f"{aba}: {dict(zip(labels, somas))} (linhas válidas {inc}, excluídas {exc})")
    print(json.dumps(out, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
